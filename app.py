from flask import Flask, render_template, request, send_file, abort
import openpyxl, os, json
from datetime import datetime
from openpyxl.styles import PatternFill

# ===== 時區設定（台灣） =====
# 優先使用 Python 3.9+ 的 zoneinfo；若環境較舊，退回 pytz
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    TAIPEI_TZ = ZoneInfo("Asia/Taipei")
except Exception:  # pragma: no cover
    import pytz
    TAIPEI_TZ = pytz.timezone("Asia/Taipei")

def now_tw():
    """回傳台灣時區的現在時間（timezone-aware）。"""
    return datetime.now(TAIPEI_TZ)

app = Flask(__name__)

# ---------------------------
# 讀取設定檔（帶基本檢查）
# ---------------------------
CONFIG_PATH = 'config.json'
if not os.path.exists(CONFIG_PATH):
    raise RuntimeError("找不到 config.json，請確認檔案已放在專案根目錄。")

with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    config = json.load(f)

TEMPLATE_PATH = config.get("input_file")
OUTPUT_FOLDER = config.get("output_folder", "outputs")

if not TEMPLATE_PATH or not os.path.exists(TEMPLATE_PATH):
    raise RuntimeError(f"找不到 Excel 範本：{TEMPLATE_PATH}，請確認 config.json 的 input_file 路徑正確且檔案存在。")

REASON_OPTIONS = [
    "休假", "曠職", "體檢", "年休返泰",
    "事假返鄉", "工傷", "病假", "待返",
    "遣返", "提前解聘", "逃跑", "調派"
]

# ✅ 固定寫死的管理員名單（無預設值，前端會有「請選擇管理員」）
MANAGER_OPTIONS = ["鄭峰源", "楊國新"]

# ---------------------------
# 工號卡控工具函式（不影響其它邏輯）
# ---------------------------
def _normalize_emp_id(v):
    """把 Excel 讀到的工號轉成純字串整數樣式（22666.0 -> '22666'）。"""
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(int(v))
    return str(v).strip()

def load_valid_emp_ids():
    """從『出勤表』(B/H/N/T/Z 欄，6~61 列)載入有效工號。"""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    try:
        if "出勤表" not in wb.sheetnames:
            raise RuntimeError("Excel 範本缺少『出勤表』工作表。")
        ws_main = wb["出勤表"]
        emp_columns = [2, 8, 14, 20, 26]  # B, H, N, T, Z
        start_row, end_row = 6, 61

        valid = set()
        for row in range(start_row, end_row + 1):
            for col in emp_columns:
                eid = _normalize_emp_id(ws_main.cell(row=row, column=col).value)
                if eid:
                    valid.add(eid)
        return valid
    finally:
        wb.close()

# 啟動時載入一次
VALID_EMP_IDS = load_valid_emp_ids()

# ---------------------------
# 原本 Excel 寫入流程（邏輯不變，時間改用台灣時區）
# ---------------------------
def update_excel(absentees, weather, manager_name=None):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_main = wb["出勤表"]
    ws_log = wb["休假調查表(新)"]

    reason_map = {emp_id.strip(): reason for emp_id, reason in absentees}
    emp_columns = [2, 8, 14, 20, 26]
    start_row, end_row = 6, 61

    fill_colors = {
        "休假": "FFFF00",  "曠職": "FF6666", "體檢": "B7DEE8", "年休返泰": "D9EAD3",
        "事假返鄉": "D0E0E3","工傷": "FFD966","病假": "C9DAF8","待返": "EAD1DC",
        "遣返": "F6B26B", "提前解聘": "A4C2F4","逃跑": "E06666","調派": "76A5AF"
    }
    count_map = {key: 0 for key in REASON_OPTIONS}

    for row in range(start_row, end_row + 1):
        for col in emp_columns:
            emp_cell = ws_main.cell(row=row, column=col)
            upper_cell = ws_main.cell(row=row, column=col + 1)
            emp_id = _normalize_emp_id(emp_cell.value)
            if len(emp_id) < 1:
                continue
            if emp_id in reason_map:
                reason = reason_map[emp_id]
                upper_cell.value = "X"
                fill_color = fill_colors.get(reason, "DDDDDD")
                emp_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if reason in count_map:
                    count_map[reason] += 1
            else:
                upper_cell.value = "V"

    # C4：yyyy年mm月dd日 星期X —— 使用台灣時區
    weekdays = ['一', '二', '三', '四', '五', '六', '日']
    today = now_tw()
    ws_main["C4"].value = today.strftime(f"%Y年%m月%d日 星期{weekdays[today.weekday()]}")

    # 天氣：P4/S4/V4 打 X
    weather_map = {"晴": "P4", "陰": "S4", "雨": "V4"}
    if weather in weather_map:
        ws_main[weather_map[weather]].value = "X"

    # 休假調查表日期 I2 —— 使用台灣時區
    ws_log["I2"].value = today.strftime("%Y年%m月%d日")

    # 休假調查表列表
    insert_row = 5
    serial_number = 1
    today_str = today.strftime("%m/%d")
    for emp_id, reason in absentees:
        ws_log.cell(row=insert_row, column=1).value = today_str
        ws_log.cell(row=insert_row, column=2).value = serial_number
        ws_log.cell(row=insert_row, column=3).value = "GC01"
        ws_log.cell(row=insert_row, column=4).value = emp_id
        ws_log.cell(row=insert_row, column=5).value = reason
        ws_log.cell(row=insert_row, column=6).value = "宿舍"
        insert_row += 1
        serial_number += 1

    # 統計欄位
    ws_main["C62"].value = count_map["體檢"]
    ws_main["C63"].value = count_map["工傷"]
    ws_main["C64"].value = count_map["遣返"]
    ws_main["K62"].value = count_map["休假"]
    ws_main["K63"].value = count_map["曠職"]
    ws_main["K64"].value = count_map["提前解聘"]
    ws_main["T62"].value = count_map["年休返泰"]
    ws_main["T63"].value = count_map["病假"]
    ws_main["T64"].value = count_map["逃跑"]
    ws_main["AA62"].value = count_map["事假返鄉"]
    ws_main["AA63"].value = count_map["待返"]
    ws_main["AA64"].value = count_map["調派"]

    # 出工人數 = D66 - 缺勤人數
    ws_main["L66"].value = int(ws_main["D66"].value or 0) - len(absentees)

    # S69：移工管理員：吳廷湘 <姓名>
    if manager_name:
        ws_main["S69"].value = f"移工管理員：吳廷湘 {manager_name}"

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    # 產出檔名 —— 使用台灣時區的日期
    output_file = f"{OUTPUT_FOLDER}/每天出工統計表_{today.strftime('%Y-%m-%d')}.xlsx"
    wb.save(output_file)
    wb.close()
    return output_file

# ---------------------------
# 路由（只加工號卡控 + 把 valid ids 傳給前端）
# ---------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        emp_ids = request.form.getlist('emp_id')
        reasons = request.form.getlist('reason')
        weather = request.form.get('weather')
        manager_name = request.form.get('manager', '').strip()

        # 工號卡控：純數字 + 必須存在
        errors = []
        cleaned_emp_ids = []
        for i, raw in enumerate(emp_ids, start=1):
            eid = (raw or "").strip()
            if not eid:
                cleaned_emp_ids.append(eid)
                continue
            if not eid.isdigit():
                errors.append(f"第 {i} 列工號需為純數字：{eid}")
            elif eid not in VALID_EMP_IDS:
                errors.append(f"第 {i} 列工號不存在於出勤表：{eid}")
            cleaned_emp_ids.append(eid)

        if errors:
            filled_rows = list(zip(emp_ids, reasons))
            # 傳 list，避免 set 序列化錯誤
            return render_template(
                'index.html',
                reasons=REASON_OPTIONS,
                managers=MANAGER_OPTIONS,
                valid_emp_ids=sorted(list(VALID_EMP_IDS)),
                errors=errors,
                filled_rows=filled_rows,
                selected_weather=weather,
                selected_manager=manager_name
            ), 400

        absentees = [(cleaned_emp_ids[i], reasons[i]) for i in range(len(cleaned_emp_ids)) if cleaned_emp_ids[i].strip()]
        result_path = update_excel(absentees, weather, manager_name)
        return send_file(result_path, as_attachment=True)

    # GET：把 valid ids 傳給前端（轉成 list）
    return render_template(
        'index.html',
        reasons=REASON_OPTIONS,
        managers=MANAGER_OPTIONS,
        valid_emp_ids=sorted(list(VALID_EMP_IDS))
    )

#（可選）健康檢查
@app.route('/health')
def health():
    return {
        "ok": True,
        "template_exists": os.path.exists(TEMPLATE_PATH),
        "ids": len(VALID_EMP_IDS),
        "now_tw": now_tw().isoformat()
    }

if __name__ == '__main__':
    # Render 建議綁定 0.0.0.0 與 PORT 環境變數
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
